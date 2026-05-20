"""双策略 tradeinfo：Excel 两页或 spread/strangle 两个 CSV。"""

import csv
import os
from typing import Any, Dict, List, Tuple

_REQUIRED = {'future', 'month', 'vol_basis', 'vol_of_combo', 'min_tick'}


def _project_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def _resolve_path(base: str, p: str) -> str:
    if os.path.isabs(p):
        return p
    if os.path.isfile(p):
        return os.path.abspath(p)
    return os.path.join(base, p)


def _parse_row(row: dict, line_no: int, source: str) -> dict:
    future = (row.get('future') or '').strip()
    month = (row.get('month') or '').strip()
    if not future or not month:
        raise ValueError(f"{source} 第{line_no}行: future 或 month 为空")
    vol_basis = float(row['vol_basis'])
    vol_of_combo = int(row['vol_of_combo'])
    min_tick = float(row['min_tick'])
    if vol_basis <= 0 or vol_of_combo <= 0 or min_tick <= 0:
        raise ValueError(f"{source} 第{line_no}行: 参数必须为正数")
    return {
        'future': future,
        'month': month,
        'vol_basis': vol_basis,
        'vol_of_combo': vol_of_combo,
        'min_tick': min_tick,
    }


def _load_csv(path: str) -> List[Dict[str, Any]]:
    items = []
    seen = set()
    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames or not _REQUIRED.issubset(set(reader.fieldnames)):
            missing = _REQUIRED - set(reader.fieldnames or [])
            raise ValueError(f"{path} 缺少列: {missing}")
        for line_no, row in enumerate(reader, start=2):
            item = _parse_row(row, line_no, path)
            key = (item['future'].lower(), item['month'])
            if key in seen:
                raise ValueError(f"{path} 第{line_no}行: 重复 {key}")
            seen.add(key)
            items.append(item)
    return items


def _load_xlsx_sheet(path: str, sheet: str) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("读取 Excel 需要: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"{path} 无工作表 '{sheet}'，现有: {wb.sheetnames}")
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else '' for c in next(rows)]
    col_map = {h: i for i, h in enumerate(header)}
    if not _REQUIRED.issubset(set(col_map)):
        raise ValueError(f"工作表 {sheet} 缺少列: {_REQUIRED - set(col_map)}")
    items = []
    seen = set()
    for line_no, row in enumerate(rows, start=2):
        if not row or all(c is None for c in row):
            continue
        data = {}
        for k in _REQUIRED:
            idx = col_map[k]
            val = row[idx] if idx < len(row) else None
            data[k] = '' if val is None else str(val).strip()
        item = _parse_row(data, line_no, f"{path}:{sheet}")
        key = (item['future'].lower(), item['month'])
        if key in seen:
            raise ValueError(f"{path}:{sheet} 第{line_no}行: 重复 {key}")
        seen.add(key)
        items.append(item)
    wb.close()
    return items


def load_dual_tradeinfo(config: dict) -> Tuple[List[dict], List[dict], List[dict]]:
    """
    返回 (spread_tradeinfo, strangle_tradeinfo, combined_for_ctp_init)。
    """
    dual = config.get('dual_strategy') or {}
    base = _project_dir()
    path = dual.get('tradeinfo_path', 'tradeinfo/tradeinfo.xlsx')
    if not os.path.isabs(path):
        path = os.path.join(base, path)

    spread_sheet = dual.get('spread_sheet', 'spread')
    strangle_sheet = dual.get('strangle_sheet', 'strangle')

    if path.lower().endswith(('.xlsx', '.xlsm')) and os.path.isfile(path):
        spread_items = _load_xlsx_sheet(path, spread_sheet)
        strangle_items = _load_xlsx_sheet(path, strangle_sheet)
    elif os.path.isdir(path):
        spread_csv = _resolve_path(base, dual.get('spread_csv', os.path.join(path, 'spread.csv')))
        strangle_csv = _resolve_path(base, dual.get('strangle_csv', os.path.join(path, 'strangle.csv')))
        spread_items = _load_csv(spread_csv)
        strangle_items = _load_csv(strangle_csv)
    else:
        trade_dir = os.path.join(base, 'tradeinfo')
        spread_csv = _resolve_path(base, dual.get('spread_csv', os.path.join(trade_dir, 'spread.csv')))
        strangle_csv = _resolve_path(base, dual.get('strangle_csv', os.path.join(trade_dir, 'strangle.csv')))
        if os.path.isfile(spread_csv) and os.path.isfile(strangle_csv):
            spread_items = _load_csv(spread_csv)
            strangle_items = _load_csv(strangle_csv)
        else:
            raise FileNotFoundError(
                f"请提供 {path}（含 {spread_sheet}/{strangle_sheet} 两页）"
                f"或 {spread_csv} + {strangle_csv}"
            )

    combined = {}
    for it in spread_items + strangle_items:
        combined[(it['future'].lower(), it['month'])] = it
    _log_tradeinfo_month_hints(spread_items, strangle_items)
    return spread_items, strangle_items, list(combined.values())


def _log_tradeinfo_month_hints(spread_items: List[dict], strangle_items: List[dict]) -> None:
    """启动前提示：同品种价差/宽跨月份或参数不一致（仅 print，logger 尚未就绪）。"""
    spread_by_future = {it['future'].lower(): it for it in spread_items}
    strangle_by_future = {it['future'].lower(): it for it in strangle_items}
    for sym in sorted(set(spread_by_future) & set(strangle_by_future)):
        s, t = spread_by_future[sym], strangle_by_future[sym]
        if s['month'] != t['month']:
            print(
                f"[TRADEINFO] {sym}: 价差 month={s['month']}, "
                f"宽跨 month={t['month']}（已支持，CTP 将订阅两个月份）"
            )
        for field in ('vol_basis', 'vol_of_combo', 'min_tick'):
            if s[field] != t[field]:
                print(
                    f"[TRADEINFO] {sym}/{s['month']}: 价差 {field}={s[field]}, "
                    f"宽跨 {field}={t[field]}（各策略独立使用，正常）"
                )
